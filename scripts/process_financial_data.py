"""
Process Excel financial model into financial_model.json.
Runs standalone — no config file required.
Usage: python scripts/process_financial_data.py
"""

from __future__ import annotations

import json
import warnings
from datetime import datetime
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).parent.parent
WORKBOOK_PATH = (
    REPO_ROOT
    / "data_extract"
    / "strategy"
    / "financial_model"
    / ("v3.6_Financial Model_FY25-2040_MASTER_v116_Bestcase.xlsx")
)
OUTPUT_DIR = REPO_ROOT / "data_extract" / "processed"
OUTPUT_FILE = OUTPUT_DIR / "financial_model.json"

YEARS = [2021, 2022, 2023, 2024, 2025, 2026, 2027, 2028, 2029, 2030]

# SAR_mn → SAR bn: multiply by 0.001
# USD_mn → SAR bn: multiply by 3.75 × 0.001
BU_CONFIG = {
    "phosphate": {"sheet": "Phos_Consolidated BU", "currency": "SAR_mn"},
    "aluminum": {"sheet": "Alum_Consolidated BU", "currency": "SAR_mn"},
    "gold": {"sheet": "BMNM_Consol", "currency": "SAR_mn"},
    "copper": {"sheet": "Kureem_Consolidated", "currency": "USD_mn"},
}

SCALE = {"SAR_mn": 0.001, "USD_mn": 3.75 * 0.001}


def _find_coords(ws) -> tuple[int | None, int | None, int | None]:
    """Returns (col_2025, ebitda_row, revenue_row). All 1-indexed."""
    col_2025 = ebitda_row = revenue_row = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=200, values_only=True), 1):
        vals = list(row[:20])
        if col_2025 is None:
            for j, v in enumerate(vals, 1):
                if v == 2025 or (
                    hasattr(v, "year") and getattr(v, "year", None) == 2025
                ):
                    col_2025 = j
                    break
        if ebitda_row is None and any(
            isinstance(v, str) and "EBITDA" in v for v in vals
        ):
            ebitda_row = i
        if revenue_row is None and any(
            isinstance(v, str) and v.strip() == "Revenue" for v in vals
        ):
            revenue_row = i
        if col_2025 and ebitda_row and revenue_row:
            break
    return col_2025, ebitda_row, revenue_row


def _extract_metric(
    ws_values, ws_formulas, row: int, col_2025: int, scale: float
) -> dict:
    values: dict[str, float] = {}
    formula_sample: str | None = None
    for year in YEARS:
        col = col_2025 + (year - 2025)
        if col < 1:
            continue
        raw = ws_values.cell(row=row, column=col).value
        if isinstance(raw, (int, float)) and raw is not None:
            values[str(year)] = round(raw * scale, 3)
            if year == 2025:
                f = ws_formulas.cell(row=row, column=col).value
                formula_sample = str(f) if f else None
    return {"values": values, "formula_sample": formula_sample, "source_row": row}


def _process_bu(bu_code: str, cfg: dict, ws_v, ws_f) -> dict | None:
    scale = SCALE[cfg["currency"]]
    col_2025, ebitda_row, revenue_row = _find_coords(ws_v)
    if not all([col_2025, ebitda_row, revenue_row]):
        warnings.warn(f"[{bu_code}] coord discovery failed — skipping")
        return None
    currency_note = (
        "SAR billions"
        if cfg["currency"] == "SAR_mn"
        else "SAR billions (converted from USD)"
    )
    return {
        "sheet": cfg["sheet"],
        "currency_note": currency_note,
        "metrics": {
            "ebitda_sar_bn": _extract_metric(ws_v, ws_f, ebitda_row, col_2025, scale),
            "revenue_sar_bn": _extract_metric(ws_v, ws_f, revenue_row, col_2025, scale),
        },
    }


def main() -> None:
    lock = WORKBOOK_PATH.parent / f"~${WORKBOOK_PATH.name}"
    if lock.exists():
        print(
            f"WARNING: Excel lock file detected ({lock.name}). Close the workbook for accurate values."
        )

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    wb_values = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True, read_only=True)
    wb_formulas = openpyxl.load_workbook(WORKBOOK_PATH, data_only=False, read_only=True)

    bus: dict = {}
    for bu_code, cfg in BU_CONFIG.items():
        if cfg["sheet"] not in wb_values.sheetnames:
            warnings.warn(f"[{bu_code}] sheet '{cfg['sheet']}' not found — skipping")
            continue
        result = _process_bu(
            bu_code,
            cfg,
            wb_values[cfg["sheet"]],
            wb_formulas[cfg["sheet"]],
        )
        if result:
            bus[bu_code] = result

    wb_values.close()
    wb_formulas.close()

    payload = {
        "schema_version": "1.0",
        "processed_at": datetime.utcnow().isoformat(),
        "source": WORKBOOK_PATH.name,
        "bus": bus,
    }
    OUTPUT_FILE.write_text(json.dumps(payload, indent=2))
    print(f"Processed {len(bus)} BUs → {OUTPUT_FILE.relative_to(REPO_ROOT)}")


if __name__ == "__main__":
    main()
