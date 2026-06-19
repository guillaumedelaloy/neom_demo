#!/usr/bin/env python3
"""Build a workbook intelligence index for the Ma'aden financial model.

Usage:
    cd <repo_root> && .venv/bin/python scripts/build_workbook_index.py

Outputs data_extract/processed/workbook_index.json — load once at agent startup
via LazyExcelLoader.index so describe_workbook() can return the full sheet
structure without any live Excel reads.
"""

from __future__ import annotations

import datetime
import json
import logging
import sys
from pathlib import Path

import openpyxl

EXCEL_PATH = (
    Path(__file__).parent.parent
    / "data_extract/strategy/financial_model"
    / "v3.6_Financial Model_FY25-2040_MASTER_v116_Bestcase.xlsx"
)
INDEX_PATH = Path(__file__).parent.parent / "data_extract/processed/workbook_index.json"

# Key sheets retain row_labels for backward-compatible describe_workbook output
_KEY_SHEETS = frozenset({
    "Consolidated Group",
    "Corporate",
    "Phosphate",
    "Aluminium",
    "Gold",
    "Phos_Consolidated BU",
    "Alum_Consolidated BU",
    "BMNM_Consol",
    "Kureem_Consolidated",
    "ARGOS_Consolidated",
})

_YEAR_MIN, _YEAR_MAX = 2020, 2045

BU_ALIASES: dict[str, str] = {
    "phosphate":    "Phos_Consolidated BU",
    "phos":         "Phos_Consolidated BU",
    "aluminum":     "Alum_Consolidated BU",
    "aluminium":    "Alum_Consolidated BU",
    "alum":         "Alum_Consolidated BU",
    "gold":         "BMNM_Consol",
    "bmnm":         "BMNM_Consol",
    "copper_kureem":"Kureem_Consolidated",
    "kureem":       "Kureem_Consolidated",
    "copper_argos": "ARGOS_Consolidated",
    "argos":        "ARGOS_Consolidated",
    "group":        "Consolidated Group",
    "corporate":    "Corporate",
}

_log = logging.getLogger(__name__)


def _longest_consecutive_run(years: list[int]) -> list[int]:
    """Return the longest run of consecutive integers from a sorted list."""
    if not years:
        return []
    best: list[int] = [years[0]]
    current: list[int] = [years[0]]
    for y in years[1:]:
        if y == current[-1] + 1:
            current.append(y)
        else:
            current = [y]
        if len(current) > len(best):
            best = list(current)
    return best


def _detect_year_headers(ws) -> tuple[int | None, dict[str, int]]:
    """Scan rows 0–5 (0-indexed) for ≥3 consecutive years in 2020–2045.

    Returns (0-based row index, {year_str: 0-based col index}) or (None, {}).
    """
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=6, min_col=1, max_col=41, values_only=True)
    ):
        years: dict[int, int] = {}
        for col_idx, val in enumerate(row):
            yr: int | None = None
            if isinstance(val, int) and _YEAR_MIN <= val <= _YEAR_MAX:
                yr = val
            elif isinstance(val, datetime.date) and _YEAR_MIN <= val.year <= _YEAR_MAX:
                yr = val.year
            if yr is not None and yr not in years:  # first occurrence wins
                years[yr] = col_idx  # 0-based
        run = _longest_consecutive_run(sorted(years))
        if len(run) >= 3:
            return (row_idx, {str(y): years[y] for y in run})
    return (None, {})


def _extract_row_labels(ws, max_row: int) -> list[dict]:
    """Extract non-empty string cells from col A (col B fallback), capped at 60."""
    labels: list[dict] = []
    limit = min(max_row, 200)
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=limit, min_col=1, max_col=2, values_only=True)
    ):
        val_a = row[0] if len(row) > 0 else None
        val_b = row[1] if len(row) > 1 else None
        if isinstance(val_a, str) and val_a.strip():
            labels.append({"row": row_idx, "col": 0, "value": val_a.strip()})
        elif isinstance(val_b, str) and val_b.strip():
            labels.append({"row": row_idx, "col": 1, "value": val_b.strip()})
        if len(labels) >= 60:
            break
    return labels


def _extract_text_tokens(ws) -> list[str]:
    """Scan all cells; collect unique strings 2–120 chars, capped at 300.

    Sort ascending by length so short, label-like strings rank first.
    """
    tokens: set[str] = set()
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if isinstance(cell, str) and cell.strip():
                cleaned = cell.strip()
                if 2 <= len(cleaned) <= 120:
                    tokens.add(cleaned)
    # Prefer shorter (more label-like) strings when capping
    return sorted(tokens, key=len)[:300]


def build_index(excel_path: Path, index_path: Path) -> int:
    """Build and write the workbook index. Returns 0 on success, 1 on error."""
    if not excel_path.exists():
        _log.error("Excel file not found: %s", excel_path)
        return 1
    _log.info("Opening workbook (read-only, data-only): %s", excel_path.name)
    try:
        wb = openpyxl.load_workbook(str(excel_path), read_only=True, data_only=True)
    except Exception as exc:
        _log.error("Failed to open workbook: %s", exc)
        return 1

    total_sheets = len(wb.sheetnames)
    sheets: dict = {}

    for name in wb.sheetnames:
        ws = wb[name]
        dims = [ws.max_row or 0, ws.max_column or 0]
        year_header_row, year_cols = _detect_year_headers(ws)
        text_tokens = _extract_text_tokens(ws)
        entry: dict = {
            "dims": dims,
            "year_header_row": year_header_row,
            "year_cols": year_cols,
            "text_tokens": text_tokens,
        }
        if name in _KEY_SHEETS:
            entry["row_labels"] = _extract_row_labels(ws, dims[0])
        sheets[name] = entry
        _log.info("[indexed] %s (%dr × %dc, %d tokens)", name, dims[0], dims[1], len(text_tokens))

    wb.close()

    index = {
        "meta": {
            "source": excel_path.name,
            "built_at": datetime.datetime.now(datetime.timezone.utc).isoformat(),
            "total_sheets": total_sheets,
            "key_sheets_count": len(_KEY_SHEETS),
        },
        "bu_aliases": BU_ALIASES,
        "sheets": sheets,
    }

    index_path.parent.mkdir(parents=True, exist_ok=True)
    content = json.dumps(index, ensure_ascii=False, indent=2)
    size_kb = len(content.encode("utf-8")) / 1024
    if size_kb > 1500:
        _log.warning("Index size %.1f KB approaching 2 MB limit", size_kb)
    if size_kb > 2048:
        _log.error("Index size %.1f KB exceeds 2 MB limit — aborting", size_kb)
        return 1

    index_path.write_text(content, encoding="utf-8")
    print(f"Indexed {total_sheets}/{total_sheets} sheets → workbook_index.json ({size_kb:.1f} KB)")
    return 0


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(message)s")
    sys.exit(build_index(EXCEL_PATH, INDEX_PATH))
